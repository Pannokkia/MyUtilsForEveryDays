import openpyxl as xls
from openpyxl.styles import colors, fonts
from openpyxl.styles import Color

def get_max_rows(wrkbk, sheet_name = 'Sheet1') -> int:
    """
        Count the maximum number of occupied rows

    Arguments: 
        
      excel_filename: excel file to process
      sheet_name: exchel sheet name to check.If sheet name not set, default is Sheet1

    Returns:
        int: Return max number occupied rows
    """

    try:
        wrkbk.active = wrkbk[sheet_name]
        return wrkbk[sheet_name].max_row
    except KeyError as e:
        print("Error:", e)
        return -1
    except PermissionError as e:
        print("Error:", e)
        return -1
    except TypeError as e:
        print("Error:", e)
        return -1

def get_max_cols(wrkbk, sheet_name = 'Sheet1') -> int:
    """
        Count the maximum number of occupied columns

    Arguments: 
        
        excel_filename: excel file to process
        sheet_name: exchel sheet name to check.If sheet name not set, default is Sheet1

    Returns:
        int: Return max number occupied columns
    """
    try:
        wrkbk.active = wrkbk[sheet_name]
        return wrkbk[sheet_name].max_column
    except KeyError as e:
        print("Error:", e)
        return -1
    except PermissionError as e:
        print("Error:", e)
        return -1
    except TypeError as e:
        print("Error:", e)
        return -1

def rename_excel_sheet(wrkbk, old_sheet_name, new_sheet_name) -> None:
    """
        Rename Sheet on excel file

    Arguments: 

        wrkbk: excel workbook
        old_excel_sheet_name: old name of ecel sheet
        new_sheet_name: new name of excel sheet

    Returns:
        None
    """
    try:
        sh = wrkbk[old_sheet_name]
        sh.title = new_sheet_name
    except KeyError as e:
        print("Error:", e)
    except PermissionError as e:
        print("Error:", e)

def remove_excel_sheet(wrkbk, sheet_name_to_remove) -> None:
    """
        Remove Sheet from excel file

    Arguments: 

        wrkbk: excel workbook
        sheet_name_to_remove: sheet to remove from excel file

    Returns:
        None
    """
    try:
        wrkbk.remove(wrkbk[sheet_name_to_remove])
    except KeyError as e:
        print("Error:", e)
    except PermissionError as e:
        print("Error:", e)

def change_style_cells(wrkbk, cells,  sheet_name = 'Sheet1',  bold = False, italic = False, strike = False, underline='none', size = 10, name='Century Gothic') -> None:
    """
    Sets the styles of  cells in an excel file 

    Arguments: 

        wrkbk: excel workbook
        cells = dictionary that contains CELLS AND COLOR of cells for which to change the style
        sheet_name = excel sheet where to set the new font style
        b = Bold
        i = Italic
        strike = strike
        umderline =  none, doubleAccounting, singleAccounting, double, single

    Returns:
        None
    """
    try:
        wrkbk.active = wrkbk[sheet_name]
        ws = wrkbk.active
        for key, value in cells.items():
            ws[key].font = fonts.Font(b = bold, i = italic, strike = strike, u = underline,color=value, size=size, name=name)
    except KeyError as e:
        print("Error:", e)
    except PermissionError as e:
        print("Error:", e)
    except IndexError as e:
        print("Error:", e)
    except ValueError as e:
        print("Error:", e)
      